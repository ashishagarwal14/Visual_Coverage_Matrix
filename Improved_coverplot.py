from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
from scipy.interpolate import make_interp_spline
import openpyxl
import glob
import xlwt
import easygui as eg
from xlwt import Workbook
import tkinter as tk
from tkinter import Frame, Label, Entry, Button
import tkinter.scrolledtext as tkst
import tkinter.messagebox
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
from tkinter.messagebox import showerror

# ---------End of imports

# starting of piller
start_point = 0
root = tk.Tk()
# importing from piller library
list_dir = glob.glob(r".\*.xlsx")

# data = pd.read_excel(r'.\datasheet123.xlsx')
# datap = pd.read_excel(r'.\pillerinfo.xlsx')
# df = pd.DataFrame(data)
# df1 = pd.DataFrame(datap)
# n0 = df.shape
# n01 = df1.shape
# arr0 = df.to_numpy()
# arr01 = df1.to_numpy()
# arr = arr0
# arr1 = arr01
# n = n0
# n1 = n01
# matplotlib parameters
norm = plt.Normalize(1, 4)
cmap = plt.cm.RdYlGn

# getting screen's height in pixels
height = root.winfo_screenheight()

# getting screen's width in pixels
width = root.winfo_screenwidth()
# print("\n width x height = %d x %d (in pixels)\n" % (width, height))

fig = plt.Figure(figsize=(12, 6))
ax = fig.add_subplot(111)

# global element of annot
annot = ax.annotate("", xy=(0, 0), xytext=(-65, 20), textcoords="offset points",
                    bbox=dict(boxstyle="round", fc="w"),
                    arrowprops=dict(arrowstyle="->", color='black'))

# globally defined rows and columns and other parameters of data
# p_row = len(arr1)
# p_col = len(arr1[0])
# d_row = len(arr)
# d_col = len(arr[0])

lines = []
ani_line = []
wo = []


class Window(Frame):

    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master = master
        self.init_window()

    def showme(self):
        for i in range(0, len(ani_line), 1):
            ani_line[i].remove()
        ani_line.clear()
        var = str(self.variable.get())
        # print(var)
        self.animate(var)
        fig.canvas.draw()

    def plot(self):
        dataset = self.excelvar.get()
        library = self.libraryvar.get()
        for i in range(0, len(list_dir), 1):
            list = list_dir[i].split("\\")
            if (list[1] == dataset):
                self.data = pd.read_excel(list_dir[i])
                # print(self.data)
                self.dfn = pd.DataFrame(self.data)
                self.n = self.dfn.shape
                arr = self.dfn.to_numpy()
                self.arr=arr
                # print(len(arr))
                # print(list_dir[i])
                d_row = len(arr)
                d_col = len(arr[0])
            elif (list[1] == library):
                # print(list_dir[i])
                self.datap1 = pd.read_excel(list_dir[i], sheet_name='Sheet1')
                self.datap2 = pd.read_excel(list_dir[i], sheet_name='Sheet2')
                self.df1 = pd.DataFrame(self.datap1)
                n1 = self.df1.shape
                arr1 = self.df1.to_numpy()
                self.arr1=arr1
                # print(self.arr1)
                self.df2 = pd.DataFrame(self.datap2)
                n2 = self.df2.shape
                arr2 = self.df2.to_numpy()
                l_row = len(arr2)
                l_col = len(arr2[0])
        self.data = [[0 for x in range(l_col)] for y in range(d_row)]
        for i in range(0, d_row, 1):
            str_data = ""
            for f in range(0, d_col, 1):
                str_data = str_data +" " +str(arr[i][f])
            str_data=str_data.lower()
            for j in range(0, l_col, 1):
                self.data[i][j]=""
                for k in range(0, l_row, 1):
                    library_s = str(arr2[k][j]).split(",")
                    flag=1
                    for l in range(0, len(library_s), 1):
                        if not library_s[l]=='N':
                            library_s[l]=library_s[l].lower()
                            if library_s[l] in str_data:
                                avail = k + 1
                                flag=0
                            else:
                                if flag==1:
                                    avail='N'
                        else:
                            avail='N'
                    avail=str(avail)
                    if k==0:
                        self.data[i][j]=avail
                    else:
                        self.data[i][j]=self.data[i][j]+","+avail
        # print(self.data)
        try:
            self.line_piller.remove()
        except Exception:
            print("Drawing pillers")
        self.cut_list = np.zeros((len(arr1), len(arr1[0])))
        for i in range(0, d_row, 1):
            self.rows.append(i)
        # for i in range(0,p_row,1):
        #     self.cut_list.append(arr1[i])
        self.tc_num = d_row
        self.log_arr=self.data
        self.pillers()
        self.Refresh()
    def Refresh(self):
        arr=self.data
        d_row,d_col=len(arr),len(arr[0])
        arr1=self.arr1
        p_row,p_col=len(arr1),len(arr1[0])
        self.log = self.arr
        self.log_arr = self.data
        self.and_count = 0
        self.or_count = 0
        for i in range(0, len(ani_line), 1):
            ani_line[i].remove()
        ani_line.clear()
        self.texttotaltc.delete("0", tk.END)
        for i in range(0, d_row, 1):
            self.rows[i] = i
        s = d_row
        self.texttotaltc.insert(0, s)
        self.logArea.delete("1.0", tk.END)
        # print(arr)
        self.full_Cov(arr)
        for i in range(len(self.log) - 1, -1, -1):
            self.logArea.insert(1.0, str(self.log[i][1]) + '\n')
            self.logArea.insert(1.0, str(self.log[i][0]) + ' : ')
            self.logArea.insert(1.0, 'TC NO: ' + str(i + 1) + " : ")
        self.w['menu'].delete(0, 'end')
        self.new_choices = []
        self.new_choices.append('________________________Select a data line__________________________________')
        # Insert list of new options (tk._setit hooks them up to var)
        for i in range(0, len(self.log), 1):
            self.new_choices.append(str(self.log[i][0]) + ":" + str(self.log[i][1]))
        # new_choices = ('one', 'two', 'three')
        self.variable.set(self.new_choices[0])
        # self.w = tk.OptionMenu(self, self.variable, *new_choices)
        for choice in self.new_choices:
            self.w['menu'].add_command(label=choice, command=tk._setit(self.variable, choice))
        fig.canvas.draw()

    def download(self):
        self.load_file()

    def or_selected(self, var):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        kk = 0
        list = []
        flag = 0
        sel_log=[]
        sel_list = []
        self.or_row = self.rows
        for i in range(0, d_row, 1):
            self.or_row.append(i)
        for i in range(0, p_col, 1):
            for j in range(0, p_row, 1):
                if (var == arr1[j][i]):
                    for k in range(1, len(self.or_row) + 1, 1):
                        if not self.or_row[k - 1] == -1:
                            stri = arr[self.or_row[k - 1]][i + start_point]
                            list = stri.split(",")
                            flag = 0
                            for l in range(0, len(list), 1):
                                if (str(j + 1) == (list[l])):
                                    flag = 1
                                    break
                            if flag == 0:
                                self.or_row[k - 1] = -1

        flag = False
        for i in range(0, len(self.or_row), 1):
            if not self.or_row[i] == -1 or self.prev_or_row[i] == -1:
                sel_list.append(arr[i])
                sel_log.append(self.arr[i])
            else:
                flag = True
        # print(sel_list)
        self.tc_num = len(sel_list)
        self.log = sel_log
        self.log_arr = sel_list
        self.rows = self.or_row
        if len(sel_list) == 0:
            print("nothing to draw")
        else:
            self.full_Cov(sel_list)
        # self.draw_cov(sel_list)

    def selected_cov(self, var):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        kk = 0
        list = []
        flag = 0
        sel_list = []
        sel_log=[]
        for i in range(0, p_col, 1):
            for j in range(0, p_row, 1):
                if (var == arr1[j][i]):
                    for k in range(1, len(self.rows) + 1, 1):
                        if not self.rows[k - 1] == -1:
                            stri = arr[self.rows[k - 1]][i + start_point]
                            list = stri.split(",")
                            flag = 0
                            for l in range(0, len(list), 1):
                                if (str(j + 1) == (list[l])):
                                    flag = 1
                                    break
                            if flag == 0:
                                self.rows[k - 1] = -1

        flag = False
        for i in range(0, len(self.rows), 1):
            if not self.rows[i] == -1:
                sel_list.append(arr[self.rows[i]])
                sel_log.append(self.arr[self.rows[i]])
            else:
                flag = True
        # print(sel_list)
        self.tc_num = len(sel_list)
        self.log = sel_log
        self.log_arr=sel_list
        if len(sel_list) == 0:
            print("nothing to draw")
        else:
            self.full_Cov(sel_list)
        # self.draw_cov(sel_list)

    def animate(self, var):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        var = str(var)
        my_list = var.split(":")
        # print(my_list)
        j = -1
        for i in range(0, len(self.log), 1):
            if str(self.log[i][0]) == str(my_list[0]):
                j = i
                # print(j)
        if j >= 0:
            col = len(arr[0])
            list = []
            # j=self.index
            x1 = []
            y1 = []
            for i in range(start_point, col, 1):
                stri = str(self.log_arr[j][i])
                list = stri.split(",")
                for k in range(0, len(list), 1):
                    if not (list[k] == 'N'):
                        y1.append(i - start_point + 1)
                        x1.append(int(list[k]))
            self.animation_line, = ax.plot(y1, x1, linestyle='dashdot', linewidth=3, marker='.', markerfacecolor='none',
                                           markersize=12,arrowprops=dict(arrowstyle="->", color=color),
                                           label='tc1')
            ani_line.append(self.animation_line)
            return j
            # self.line.set_ydata(self.A*np.sin(self.x+self.v*i)++self.R)  # update the data
            # self.full_Cov()
            # fig.canvas.mpl_connect("motion_notify_event", self.hover)
        else:
            return 0

    def update_annot_points(self, ind):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        # print('update')
        # line=self.pillers()
        x, y = self.line_piller.get_data()
        # pos = ax.get_offsets()[ind["ind"][0]]
        # annot.xy = pos
        annot.xy = (x[ind["ind"][0]], y[ind["ind"][0]])
        # print(x[ind["ind"][0]], y[ind["ind"][0]])
        text = "{}, {}".format(" ".join(str(arr1[y[ind["ind"][0]] - 1][x[ind["ind"][0]] - 1])),
                               " ".join(str(int(self.cut_list[y[ind["ind"][0]] - 1][x[ind["ind"][0]] - 1]))))
        annot.set_text(text)
        annot.get_bbox_patch().set_alpha(0.9)

    def effect(self):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        # print(self.check.get())
        if self.check.get() == 0:
            for i in range(0, len(ani_line), 1):
                ani_line[i].remove()
            ani_line.clear()
            # self.ani = animation.FuncAnimation(fig, self.animate, interval=1000, blit=False)
            # self.animate(1)

            print('line cleared')

    def update_annot_line(self, line, idx):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        # ani_line[0].remove()
        for i in range(0, len(ani_line), 1):
            ani_line[i].remove()
        ani_line.clear()
        posx, posy = [line.get_xdata()[idx], line.get_ydata()[idx]]

        self.index = int(line.get_label())
        index = self.index
        self.variable.set(self.new_choices[index + 1])
        self.animate(self.new_choices[index+1])
        # self.ani = animation.FuncAnimation(fig, self.animate, interval=100, blit=False)
        # text = f'{str(self.log[index][0]) + " :" + str(self.log[index][1])}'
        # ar=self.log
        # row = len(ar)
        # col = len(ar[0])
        # y1 = []
        # list = []
        # j=index
        # self.anno=[]
        # for i in range(start_point, col, 1):
        #     stri = str(ar[j][i])
        #     list = stri.split(",")
        #     for k in range(0, len(list), 1):
        #         if not (list[k] == 'N'):
        # annot = ax.annotate("", xy=(0, 0), xytext=(-65, 20), textcoords="offset points",
        #                                  bbox=dict(boxstyle="round", fc="w"),
        #                                  arrowprops=dict(arrowstyle="->"))
        # annot.xy = (posx, posy)
        # text = "{}".format(" ".join('TC: '+str(index)+':'+str(self.log[index][0])))
        # annot.set_text(text)
        # annot.get_bbox_patch().set_alpha(0.9)
        # annot.set_visible(True)
        # self.texthover.delete("0", tk.END)
        # self.texthover.insert(0, str(self.log[index][1]))
        #             self.anno.append(anno)
        # print(line.get_label())
        # annot.get_bbox_patch().set_facecolor(cmap(norm(c[ind["ind"][0]])))

    def onpick(self, event):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        # print("123")
        for i in range(0, len(ani_line), 1):
            ani_line[i].remove()
        ani_line.clear()
        self.line_piller = event.artist
        xdata = self.line_piller.get_xdata()
        ydata = self.line_piller.get_ydata()
        ind = event.ind
        points = tuple(zip(xdata[ind], ydata[ind]))
        # print('onpick points:', points)
        # print(arr1)
        for i in range(0, len(lines), 1):
            lines[i].remove()
            flag = 0
        lines.clear()
        var = str(self.rule.get())
        if var == self.choice[0]:
            if self.and_count > 0 and self.or_count > 0:
                self.rows = self.or_row
            self.selected_cov(str(arr1[int(ydata[ind] - 1)][int(xdata[ind] - 1)]))
            self.and_count = 1 + self.and_count
        else:
            if self.or_count == 0:
                if self.and_count == 0:
                    self.prev_or_row = []
                    for i in range(0, d_row, 1):
                        self.prev_or_row.append(-1)
                elif self.and_count > 0:
                    self.prev_or_row = self.rows

            else:
                self.prev_or_row = self.or_row
            self.or_selected(str(arr1[int(ydata[ind] - 1)][int(xdata[ind] - 1)]))
            self.or_count = self.or_count + 1
        # print(str(arr1[int(ydata[ind]-1)][int(xdata[ind]-1)]))
        s = str(self.tc_num)
        self.texttotaltc.delete("0", tk.END)
        self.texttotaltc.insert(0, s)
        self.logArea.delete("1.0", tk.END)
        for i in range(len(self.log) - 1, -1, -1):
            self.logArea.insert(1.0, str(self.log[i][1]) + '\n')
            self.logArea.insert(1.0, str(self.log[i][0]) + ' : ')
            self.logArea.insert(1.0, 'TC NO: ' + str(i + 1) + " : ")
        # var.set('')
        self.w['menu'].delete(0, 'end')
        self.new_choices = []
        self.new_choices.append('________________________Select a data line__________________________________')
        # Insert list of new options (tk._setit hooks them up to var)
        for i in range(0, len(self.log), 1):
            self.new_choices.append(str(self.log[i][0]) + ":" + str(self.log[i][1]))
        # new_choices = ('one', 'two', 'three')
        self.variable.set(self.new_choices[0])
        # self.w = tk.OptionMenu(self, self.variable, *new_choices)
        for choice in self.new_choices:
            self.w['menu'].add_command(label=choice, command=tk._setit(self.variable, choice))
        fig.canvas.draw()

    def hover(self, event):
        try:
            arr = self.data
            d_row, d_col = len(arr), len(arr[0])
            arr1 = self.arr1
            p_row, p_col = len(arr1), len(arr1[0])
            # annot = ax.annotate("", xy=(0, 0), xytext=(-20, 20), textcoords="offset points",
            #                     bbox=dict(boxstyle="round", fc="w"),
            #                     arrowprops=dict(arrowstyle="->"))
            # annot.set_visible(False)
            # print('hover'+str(event))
            vis = annot.get_visible()
            if event.inaxes == ax and self.check.get() == 0:
                cont, ind = self.line_piller.contains(event)
                if cont:
                    self.update_annot_points(ind)
                    annot.set_visible(True)
                    fig.canvas.draw_idle()
                else:
                    if vis:
                        annot.set_visible(False)
                        fig.canvas.draw_idle()
            if event.inaxes == ax and self.check.get():
                for line in lines:
                    cont, ind = line.contains(event)
                    if cont:
                        self.update_annot_line(line, ind['ind'][0])
                        annot.set_visible(True)

                        fig.canvas.draw_idle()
                    else:
                        if vis:
                            annot.set_visible(False)
                            # for i in range(0, len(self.anno), 1):
                            #     self.anno[i].set_visible(False)
                            # annot.set_visible(False)
                            fig.canvas.draw_idle()
        except Exception:
            flag=0

    '''def pillers(self):
        try:
            arr = self.data
            d_row, d_col = len(arr), len(arr[0])
            arr1 = self.arr1
            p_row, p_col = len(arr1), len(arr1[0])
            x1 = []
            for i in range(1, p_row + 1, 1):
                x1.append(i)
            y1 = []
            for i in range(1, p_col + 1, 1):
                j = 1
                while (j <= p_row):
                    if not i == p_col: 
                        x1.append(x1[j - 1])
                    y1.append(i)
                    if not(str(arr1[j-1][i-1])=="N" or len(str(arr1[j-1][i-1]))==0):
                        annots = ax.annotate("", xy=(0, 0), xytext=(-30, -15), textcoords="offset points", fontsize=5,rotation=45)
                        annots.xy = (i, j)
                        text = "{}".format(" ".join(str(arr1[j - 1][i - 1])))
                        annots.set_text(text)
                        # annot.get_bbox_patch().set_alpha(0.9)
                        annots.set_visible(True)
                        j = j + 1
            print(str(x1))
            print(str(y1))
            self.line_piller, = ax.plot(y1, x1, color='black', linestyle='None', linewidth=1, marker='o',
                                        markerfacecolor='red',
                                        markersize=5, picker=5)
            fig.canvas.draw()
        except Exception:
            print("error while drawing the pillers")'''
            
    def pillers(self):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        x1 = []
        y1=[]
        for j in range(1, p_col+1 ,1):
            for i in range(1, p_row+1, 1):
                if not(str(arr1[i-1][j-1])=="N"):
                    print(arr1[i-1][j-1])
                    x1.append(i)
                    y1.append(j)
                    annots = ax.annotate("", xy=(0, 0), xytext=(-22, -25), textcoords="offset points", fontsize=6)
                    annots.xy = (j, i)
                    #text = "{}".format(" ".join("taking test"+"\n"+"what"))
                    text = "{}".format(" ".join(str(arr1[i - 1][j - 1])[0:10]+"\n"+str(arr1[i-1][j-1])[10:20]+"\n"+str(arr1[i-1][j-1])[20:30]))
                    annots.set_text(text)
                    # annot.get_bbox_patch().set_alpha(0.9)
                    annots.set_visible(True)
        self.line_piller, = ax.plot(y1, x1, color='black', linestyle='None', linewidth=1, marker='o',
                                        markerfacecolor='red',
                                        markersize=5, picker=5)
        ax.set_ylim(bottom=0.5)
        print(str(x1))
        #print(str(arr1[1][2]))
        print(str(y1))
        fig.canvas.draw()

    def full_Cov(self, ar):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        row = len(ar)
        col = len(ar[0])
        y1 = []
        list = []
        for j in range(0, row, 1):
            x1 = []
            y1 = []
            for i in range(start_point, col, 1):
                stri = str(ar[j][i])
                list = stri.split(",")
                for k in range(0, len(list), 1):
                    if not (list[k] == 'N'):
                        y1.append(int(i - start_point + 1))
                        x1.append(int(list[k]))
            Y_ = np.linspace(y1[0],y1[len(y1)-1], 5000)
            print(str(Y_))
            X_Y_Spline = make_interp_spline(y1, x1)
            X_ = X_Y_Spline(Y_)
            self.line, = ax.plot(Y_, X_, linestyle='dashed', linewidth=1, marker='.', markerfacecolor='blue',
                                 markersize=1, label="line1")
            lines.append(self.line)
            #
            #    plt.text(y1[j], 6, 'This text starts at point'
        # self.lines
        self.colorcode(ar)

    def colorcode(self, ar):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        row = len(ar)
        col = len(ar[0])
        list = []
        # print(self.tc_num)
        for i in range(1, p_col + 1, 1):
            for j in range(1, p_row + 1, 1):
                num = 0
                for k in range(0, row, 1):
                    stri = str(ar[k][i + start_point - 1])
                    list = stri.split(",")
                    for l in range(0, len(list), 1):
                        if str(j) == list[l]:
                            num = num + 1
                            # print(num)
                            break
                self.cut_list[j - 1][i - 1] = num
                if num <= int(row) / 5 and not num==0:
                    ax.plot(i, j, linestyle='dashed', linewidth=1, marker='.', markerfacecolor='yellow',
                            markersize=15,
                            label='tc1')

                if num <= int(row) / 10 and not num==0:
                    ax.plot(i, j, linestyle='dashed', linewidth=1, marker='.', markerfacecolor='red',
                            markersize=15,
                            label='tc1')
                if num == 0:
                    print("no line touching this point")

    def load_file(self):
        arr = self.data
        d_row, d_col = len(arr), len(arr[0])
        arr1 = self.arr1
        p_row, p_col = len(arr1), len(arr1[0])
        # Call a Workbook() function of openpyxl
        # to create a new blank Workbook object
        wb = openpyxl.Workbook()

        # Get workbook active sheet
        # from the active attribute
        sheet = wb.active

        # Cell objects also have row, column
        # and coordinate attributes that provide
        # location information for the cell.

        # Note: The first row or column integer
        # is 1, not 0. Cell object is created by
        # using sheet object's cell() method.
        for i in range(0, 2, 1):
            for j in range(0, len(self.log), 1):
                c1 = sheet.cell(row=j + 1, column=i + 1)
                c1.value = self.log[j][i]
        document = wb
        # insert_sheet = document.add_sheet('sheet1')
        # cambio = insert_sheet.set_portrait(False)
        extension = ["*.xlsx"]
        filesave = eg.filesavebox(msg="save file",
                                  title="", default='coverage_assistant',
                                  filetypes=extension)
        if (filesave is not None) and (len(filesave) != 0):
            document.save(str(filesave) + ".xlsx")
            print("Download completed")
        else:
            print("Download failed")
            return False

    def init_window(self):

        self.master.title("COVERPLOT")
        self.pack(fill='both', expand=1)

        # intializations
        self.and_count = 0
        self.or_count = 0
        self.rows = []
        self.cut_list = np.zeros((5, 5))
        # ani_line_a=ax.plot(1, 1, linestyle='dashdot', linewidth=2, marker='.', markerfacecolor='none',
        #         markersize=12,
        #         label='tc1')
        # ani_line.append(ani_line_a)

        self.tc_num = 0

        # self.labelmessage = Label(self,text="Select specific nodes to see related coverage",width=40)
        # self.labelmessage.grid(row=0,column=2)
        self.labeltotaltc = Label(self, text="Total data lines on plot", width=40)
        self.labeltotaltc.grid(row=0, column=0)
        self.texttotaltc = Entry(self, width=10)
        self.texttotaltc.grid(row=1, column=0)
        self.texttotaltc.insert(0, 0)
        self.labelexcel = Label(self, text="Download the displayed data into excel files", width=40)
        # self.labelexcel.grid(row=5, column=6)
        # self.textexcel = Entry(self, width=40)
        # self.textexcel.grid(row=5, column=6)
        # self.textexcel.insert(0, d_row)

        self.logArea = tkst.ScrolledText(self,
                                         wrap=tk.WORD,
                                         width=70,
                                         height=35,
                                         # state="disabled",
                                         name="logArea"
                                         )
        # self.logArea.grid(row=8, column=5, columnspan=4, rowspan=10, sticky=tk.W)
        self.logArea.insert('1.0', 'test case intializing: press on plot to see details')

        # self.texttcinfo = Entry(self, width=50)
        # self.texttcinfo.grid(row=5, column=1)

        self.buttonRefresh = Button(self, text="Refresh", command=self.Refresh, width=12)
        self.buttonRefresh.grid(row=4, column=4)

        self.buttondownload = Button(self, text="Download", command=self.download, width=12)
        self.buttondownload.grid(row=5, column=4)

        self.check = tk.IntVar()
        self.Button1 = tk.Checkbutton(self, text="Hover on lines",
                                      variable=self.check,
                                      onvalue=1,
                                      offvalue=0,
                                      height=2,
                                      width=10)
        self.Button1.grid(row=6, column=0,columnspan=1)
        # self.buttoneffect = Button(self, text="Effect", command=self.effect, width=12)
        # self.buttoneffect.grid(row=5, column=2)

        # tk.Label(self,text="Test content").grid(column=1, row=7)

        self.x = 20 * np.arange(0, 2 * np.pi, 0.01)  # x-array
        # self.pillers()
        # self.line , = self.ax.plot(self.x, np.sin(self.x))
        self.canvas = FigureCanvasTkAgg(fig, master=self)
        self.canvas.get_tk_widget().grid(row=8, column=0, columnspan=5, rowspan=10
                                         )
        self.variable = 0
        # for i in range(0,d_row,1):
        #     self.variable.append(0)
        self.new_choices = []
        self.new_choices.append('Select a data line to highlight on plot')
        # for i in range(0, d_row, 1):
        #     self.new_choices.append(str(arr[i][0]) + ":" + str(arr[i][1]))
        self.new_choices.append("Nothing to show")
        self.variable = tk.StringVar(self)
        self.variable.set(self.new_choices[0])
        self.w = tk.OptionMenu(self, self.variable, *self.new_choices)
        self.w.config(width= 100)
        self.w.grid(row=5, column=0, columnspan=4)
        self.rule = 0
        # for i in range(0,d_row,1):
        #     self.variable.append(0)
        self.choice = ['And Operation', 'Or Operation']
        # self.choice.append('________________________Select a data line__________________________________')

        self.rule = tk.StringVar(self)
        self.rule.set(self.choice[0])
        self.w_rule = tk.OptionMenu(self, self.rule, *self.choice)
        # wo.append(self.w)
        # self.w_rule.grid(row=5, column=0)
        self.labelhover = Label(self, text="Info of hovered data-line/Select a data line to see in below box", width=60)
        self.labelhover.grid(row=4, column=0,columnspan=2)
        self.buttonshowme = Button(self, text="ShowMe", command=self.showme, width=12)
        self.buttonshowme.grid(row=6, column=1)
        # self.texthover = Entry(self, width=100)
        # self.texthover.grid(row=22, column=0)
        # self.texthover.insert(0, 'Select hover on tc and click on effect to start')
        # self.full_Cov()
        # self.log = arr

        OPTIONS = ["Select Dataset"]
        for i in range(0, len(list_dir), 1):
            list = list_dir[i].split("\\")
            OPTIONS.append(list[1])

        self.excelvar = tk.StringVar(self)
        self.excelvar.set(OPTIONS[0])
        excel_w = tk.OptionMenu(self, self.excelvar, *OPTIONS)
        excel_w.config(width=40)
        # wo.append(excwl_w)
        excel_w.grid(row=0, column=1, columnspan=2)

        OPTIONS = ["Select Library"]
        for i in range(0, len(list_dir), 1):
            list = list_dir[i].split("\\")
            OPTIONS.append(list[1])

        self.libraryvar = tk.StringVar(self)
        self.libraryvar.set(OPTIONS[0])
        library_w = tk.OptionMenu(self, self.libraryvar, *OPTIONS)
        library_w.config(width=40)
        # wo.append(excwl_w)
        library_w.grid(row=0, column=3, columnspan=2)
        self.buttonplot = Button(self, text="Plot", command=self.plot, width=12)
        self.buttonplot.grid(row=2, column=4)

        # data = pd.read_excel(r'.\datasheet123.xlsx')
        # datap = pd.read_excel(r'.\pillerinfo.xlsx')
        # df = pd.DataFrame(data)
        # df1 = pd.DataFrame(datap)
        # n = df.shape
        # n1 = df1.shape
        # arr = df.to_numpy()
        # arr1 = df1.to_numpy()
        w,h=5,5
        self.arr = [[0 for x in range(w)] for y in range(h)]
        self.data = [[0 for x in range(w)] for y in range(h)]
        # self.ani = animation.FuncAnimation(fig, self.animate, np.arange(1, 200), interval=2000, blit=False)
        # self.pillers()
        fig.canvas.mpl_connect('pick_event', self.onpick)
        fig.canvas.mpl_connect("motion_notify_event", self.hover)
        #canvas = FigureCanvasTkAgg(fig, master=root)
        self.toolbar = NavigationToolbar2Tk(self.canvas, root)
        self.toolbar.update()
        #self.canvas.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)

        self.scrollbar = tkinter.Scrollbar(master=root)
        self.scrollbar.pack(side=tkinter.BOTTOM)
        self.scrollbar["command"] = self.canvas.get_tk_widget().xview
        self.canvas.get_tk_widget()["xscrollcommand"] = self.scrollbar.set


root.geometry("1000x1000")
app = Window(root)
tk.mainloop()
